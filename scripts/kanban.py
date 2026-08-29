#!/usr/bin/env python3
"""docs/kanban.xlsx を更新する最小 kanban。

  python3 scripts/kanban.py init                       # 初期カードで作り直す
  python3 scripts/kanban.py show                       # ボードを表示
  python3 scripts/kanban.py move S02 doing "担当官に発注"  # 列を移動(backlog/doing/review/done)
  python3 scripts/kanban.py note S02 "検収 3/5"          # メモ追記

シート "Cards" が正本、シート "Board" は Cards から毎回描き直す。
"""
import sys, datetime, pathlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PATH = pathlib.Path(__file__).resolve().parents[1] / "docs" / "kanban.xlsx"
COLS = ["backlog", "doing", "review", "done"]
COL_LABEL = {"backlog": "Backlog", "doing": "In Progress", "review": "Review(Qodo)", "done": "Done"}
FILL = {"backlog": "EEEEEE", "doing": "FFF2CC", "review": "DDEBF7", "done": "E2EFDA"}
HEAD = ["ID", "Title", "Step doc", "Status", "Owner", "Updated", "Note"]

INITIAL = [
    ("S00", "骨組み(Package / .gitignore / Qodo / 発注書 / kanban)", "docs/steps/step-00-scaffold.md"),
    ("S01", "ハーネス(TrueForge + OpenAI + Bright Data MCP + 呪文)", "docs/steps/step-01-harness.md"),
    ("S02", "索敵・特定 `aishow scan`(ContextPack / detect)", "docs/steps/step-02-scan-detect.md"),
    ("S03", "発動 `aishow cast`(貼り付け・クリップボード復元)", "docs/steps/step-03-cast.md"),
    ("S04", "詠唱 `aishow chant`(録音 → OpenAI STT)", "docs/steps/step-04-chant.md"),
    ("S05", "召喚 `aishow summon`(TrueForge API・承認 → cast)", "docs/steps/step-05-summon.md"),
    ("S06", "常駐 `Aishow.app`(メニューバー・ホットキー・承認 UI)", "docs/steps/step-06-menubar.md"),
    ("D01", "デモ動画(3 分)・README 仕上げ", "docs/idea.md#6"),
    ("D02", "提出フォーム(Q8–Q19)", "../hackathon/agent_harness_20260829/docs/submission.md"),
]

def now(): return datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

def load():
    if not PATH.exists(): init(); 
    wb = load_workbook(PATH); ws = wb["Cards"]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    return wb, rows

def save(wb, rows):
    ws = wb["Cards"]; ws.delete_rows(1, ws.max_row)
    ws.append(HEAD)
    for c in ws[1]: c.font = Font(bold=True)
    for r in rows: ws.append(r)
    for i, w in enumerate([6, 58, 40, 10, 10, 17, 60], 1): ws.column_dimensions[get_column_letter(i)].width = w
    render_board(wb, rows); wb.save(PATH)

def render_board(wb, rows):
    if "Board" in wb.sheetnames: del wb["Board"]
    ws = wb.create_sheet("Board", 0)
    for i, col in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i, value=f"{COL_LABEL[col]} ({sum(1 for r in rows if r[3]==col)})")
        c.font = Font(bold=True, size=12); c.fill = PatternFill("solid", fgColor=FILL[col]); c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = 46
        for j, r in enumerate([r for r in rows if r[3] == col], 2):
            cell = ws.cell(row=j, column=i, value=f"{r[0]}  {r[1]}\n{r[4] or ''} · {r[5]}\n{r[6] or ''}".rstrip())
            cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.fill = PatternFill("solid", fgColor=FILL[col])
            ws.row_dimensions[j].height = 58
    ws.cell(row=1, column=6, value=f"updated {now()}").font = Font(italic=True, color="888888")

def init():
    wb = Workbook(); wb.active.title = "Cards"
    rows = [[i, t, d, "backlog", "", now(), ""] for i, t, d in INITIAL]
    save(wb, rows)

def find(rows, cid):
    for r in rows:
        if r[0] == cid: return r
    sys.exit(f"no card {cid}")

def show(rows):
    for col in COLS:
        print(f"\n== {COL_LABEL[col]} ==")
        for r in rows:
            if r[3] == col: print(f"  {r[0]}  {r[1]}  [{r[4] or '-'}] {r[6] or ''}")

if __name__ == "__main__":
    a = sys.argv[1:]
    if not a or a[0] == "show":
        _, rows = load(); show(rows)
    elif a[0] == "init":
        init(); print("initialized", PATH)
    elif a[0] == "move" and len(a) >= 3 and a[2] in COLS:
        wb, rows = load(); r = find(rows, a[1]); r[3] = a[2]; r[5] = now()
        if len(a) > 3: r[6] = a[3]
        if a[2] == "doing" and not r[4]: r[4] = "agent"
        save(wb, rows); show(rows)
    elif a[0] == "note" and len(a) >= 3:
        wb, rows = load(); r = find(rows, a[1]); r[6] = a[2]; r[5] = now(); save(wb, rows); show(rows)
    elif a[0] == "owner" and len(a) >= 3:
        wb, rows = load(); r = find(rows, a[1]); r[4] = a[2]; r[5] = now(); save(wb, rows); show(rows)
    else:
        sys.exit(__doc__)
